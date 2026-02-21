import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])

import os, yaml, ssl, collections
import torch
import torch.nn as nn
import torch.optim as optim
import pandas as pd
from sys import platform
from openpyxl import Workbook, load_workbook

from src.model.PixelDA.generator_model import Generator, weights_init_normal
from src.model.PixelDA.discriminator_model import Discriminator
from src.model.PixelDA.classifier_model import Classifier
import src.model.PixelDA.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context


# ---------------- Config ----------------
if platform == "win32":
    cfg_path = "./configs/pixelda_train.yaml"
else:
    cfg_path = "/mimer/NOBACKUP/groups/snic2022-5-277/rofena/Progetto_Alessandro/configs/pixelda_train.yaml"

with open(cfg_path) as f:
    cfg = yaml.load(f, Loader=yaml.FullLoader)

exp_name = cfg["exp_name"]
model_name = cfg["model"]["model_name"]

device = torch.device(cfg["device"]["cuda_device"] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg["device"]["gpu_num_workers"]

cv = cfg["data"]["cv"]
fold_list = list(range(cv))

fold_dir = cfg["data"]["fold_dir"]
model_dir = os.path.join(cfg["data"]["model_dir"], exp_name)
report_dir = os.path.join(cfg["data"]["report_dir"], exp_name)
plot_training_dir = os.path.join(report_dir, "training")
outputs_dir = os.path.join(report_dir, "outputs")

for d in [model_dir, report_dir, plot_training_dir, outputs_dir]:
    util_general.create_dir(d)

report_file_mse  = os.path.join(report_dir, "report_mse.xlsx")
report_file_psnr = os.path.join(report_dir, "report_psnr.xlsx")
report_file_vif  = os.path.join(report_dir, "report_vif.xlsx")
report_file_ssim = os.path.join(report_dir, "report_ssim.xlsx")
report_file_acc  = os.path.join(report_dir, "report_acc.xlsx")

results_mse = collections.defaultdict(lambda: [])
results_psnr = collections.defaultdict(lambda: [])
results_vif = collections.defaultdict(lambda: [])
results_ssim = collections.defaultdict(lambda: [])
results_acc = collections.defaultdict(lambda: [])

metric_cols = []

# -------- Dataset wrapper con label regione --------
class PixelDADataset(util_data.ImgDataset):
    def __init__(self, data, cfg_data, step, do_augmentation=True, label_col="region"):
        super().__init__(data=data, cfg_data=cfg_data, step=step, do_augmentation=do_augmentation)
        self.label_col = label_col
        self.map_region = {"AB":0, "HN":1, "TH":2}

    def __getitem__(self, index):
        img_x, img_y, id_slice = super().__getitem__(index)
        row = self.data.iloc[index]
        reg = str(row[self.label_col]).upper()
        y = self.map_region[reg]
        y = torch.tensor(y, dtype=torch.long)
        return img_x, img_y, y, id_slice


for fold in fold_list:
    print(f"\n===== Fold {fold} =====")

    model_fold_dir = os.path.join(model_dir, str(fold))
    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    for d in [model_fold_dir, plot_training_fold_dir, outputs_fold_dir]:
        util_general.create_dir(d)

    metric_cols.append(f"{fold}")

    # ---- load CSV (train/val/test) ----
    fold_data = {
        step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col="id_slice")
        for step in ["train","val","test"]
    }

    datasets = {
        step: PixelDADataset(
            data=fold_data[step],
            cfg_data=cfg["data"],
            step=step,
            do_augmentation=True,
            label_col=cfg["data"].get("label_col", "region")
        )
        for step in ["train","val","test"]
    }

    data_loaders = {
        "train": torch.utils.data.DataLoader(
            datasets["train"], batch_size=cfg["data"]["batch_size"],
            shuffle=True, num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True, persistent_workers=True if num_workers>0 else False
        ),
        "val": torch.utils.data.DataLoader(
            datasets["val"], batch_size=cfg["data"]["batch_size"],
            shuffle=False, num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True, persistent_workers=True if num_workers>0 else False
        ),
        "test": torch.utils.data.DataLoader(
            datasets["test"], batch_size=1,
            shuffle=False, num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True, persistent_workers=True if num_workers>0 else False
        ),
    }

    if fold == 0:
        wb = Workbook()
        info_path = os.path.join(model_dir, "info.xlsx")
        wb.save(info_path)
        wb = load_workbook(info_path)
        sheet = wb.active
        sheet["A1"] = "exp_name"; sheet["A2"] = exp_name
        sheet["B1"] = "batch size"; sheet["B2"] = cfg["data"]["batch_size"]
        sheet["C1"] = "epochs"; sheet["C2"] = cfg["trainer"]["max_epochs"]
        sheet["D1"] = "img_dim"; sheet["D2"] = cfg["data"]["img_dim"]
        wb.save(info_path)

    # ---- models ----
    G = Generator(
        img_channels=1,
        latent_dim=cfg["trainer"].get("latent_dim", 8),
        n_residual_blocks=cfg["model"].get("n_residual_blocks", 6),
        out_activation=cfg["model"].get("out_activation", "sigmoid")
    ).to(device)
    D = Discriminator(in_channels=1).to(device)
    C = Classifier(in_channels=1, n_classes=3, img_dim=cfg["data"]["img_dim"]).to(device)

    G.apply(weights_init_normal)
    D.apply(weights_init_normal)
    C.apply(weights_init_normal)

    lr = cfg["trainer"]["optimizer"]["lr"]
    opt_GC = optim.Adam(list(G.parameters()) + list(C.parameters()), lr=lr, betas=(0.5, 0.999))
    opt_D  = optim.Adam(D.parameters(), lr=lr, betas=(0.5, 0.999))

    adv_loss  = nn.MSELoss().to(device)
    task_loss = nn.CrossEntropyLoss().to(device)

    # ---- train ----
    G, D, C, history = util_model.train_pixelda(
        G, D, C,
        data_loaders=data_loaders,
        opt_GC=opt_GC, opt_D=opt_D,
        adv_loss=adv_loss, task_loss=task_loss,
        cfg_trainer=cfg["trainer"],
        model_fold_dir=model_fold_dir,
        device=device
    )

    util_model.plot_training(history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # ---- test ----
    mse_t, psnr_t, vif_t, ssim_t, acc_t = util_model.test_4metrics_pixelda(
        G, C, loader_test=data_loaders["test"], device=device,
        outputs_dir=outputs_fold_dir, save_output=True
    )
    print(f"Fold {fold} → MSE={mse_t:.6f}, PSNR={psnr_t:.3f}, VIF={vif_t:.3f}, SSIM={ssim_t:.3f}, ACC={acc_t*100:.1f}%")

    results_mse[f"{fold} MSE"].append(mse_t)
    results_psnr[f"{fold} PSNR"].append(psnr_t)
    results_vif[f"{fold} VIF"].append(vif_t)
    results_ssim[f"{fold} SSIM"].append(ssim_t)
    results_acc[f"{fold} ACC"].append(acc_t)

    # ---- save reports ----
    def save_report(dic, cols, path, metric_name):
        df = pd.DataFrame(dic)
        df.insert(0, f"std {metric_name}", df[cols].std(axis=1))
        df.insert(0, f"mean {metric_name}", df[cols].mean(axis=1))
        df.insert(0, "model", model_name)
        df.to_excel(path, index=False)

    cols_mse  = [f"{f} MSE" for f in fold_list if f"{f} MSE" in results_mse]
    cols_psnr = [f"{f} PSNR" for f in fold_list if f"{f} PSNR" in results_psnr]
    cols_vif  = [f"{f} VIF" for f in fold_list if f"{f} VIF" in results_vif]
    cols_ssim = [f"{f} SSIM" for f in fold_list if f"{f} SSIM" in results_ssim]
    cols_acc  = [f"{f} ACC" for f in fold_list if f"{f} ACC" in results_acc]

    save_report(results_mse, cols_mse, report_file_mse, "MSE")
    save_report(results_psnr, cols_psnr, report_file_psnr, "PSNR")
    save_report(results_vif, cols_vif, report_file_vif, "VIF")
    save_report(results_ssim, cols_ssim, report_file_ssim, "SSIM")
    save_report(results_acc, cols_acc, report_file_acc, "ACC")
