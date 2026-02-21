import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])

import os, yaml, ssl, collections
import torch
import torch.nn as nn
import torch.optim as optim
import pandas as pd
from sys import platform
from openpyxl import Workbook, load_workbook
from torchmetrics import PeakSignalNoiseRatio

from src.model.Cogan.generator_model import CoupledGenerator, weights_init_normal
from src.model.Cogan.discriminator_model import CoupledDiscriminator, weights_init_normal as w_init_D
import src.model.Cogan.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data

# --------------------------------------------------
# Setup
# --------------------------------------------------
torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# ==========================
# CONFIG
# ==========================
if platform == 'win32':
    cfg_path = "./configs/cogan_train.yaml"
else:
    cfg_path = ".../configs/cogan_train.yaml"

with open(cfg_path) as f:
    cfg = yaml.load(f, Loader=yaml.FullLoader)

exp_name       = cfg['exp_name']
model_name     = cfg['model']['model_name']
cv             = cfg['data']['cv']
exp_name_excel = cfg.get('exp_name_excel', exp_name)

print(exp_name)


fold_env = os.environ.get("FOLD_IDX", None)
if fold_env is not None:
    fold_list = [int(fold_env)]
    print(f"*** Esecuzione solo per la fold {fold_list[0]} (da variabile d'ambiente FOLD_IDX) ***")
else:
    fold_list = list(range(cv))
    print(f"*** Esecuzione per tutte le fold: {fold_list} ***")

device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print("Device:", device)

# ==========================
# DIRS
# ==========================
fold_dir = os.path.join(cfg['data']['fold_dir'])

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

plot_training_dir = os.path.join(report_dir, "training")
outputs_dir       = os.path.join(report_dir, "outputs")
util_general.create_dir(plot_training_dir)
util_general.create_dir(outputs_dir)


fold_metrics_full = []

# ==========================
# LOOP FOLD
# ==========================
for fold in fold_list:
    print(f"\n========== Fold {fold} ==========")
    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)

    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    outputs_fold_dir       = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    util_general.create_dir(outputs_fold_dir)

    # ---------- DataLoader ----------
    fold_data = {
        step: pd.read_csv(os.path.join(fold_dir, str(fold), f'{step}.csv'), index_col='id_slice')
        for step in ['train', 'val', 'test']
    }

    datasets = {
        step: util_data.ImgDataset(
            data=fold_data[step],
            cfg_data=cfg['data'],
            step=step,
            do_augmentation=True
        )
        for step in ['train', 'val', 'test']
    }

    data_loaders = {
        'train': torch.utils.data.DataLoader(
            datasets['train'],
            batch_size=cfg['data']['batch_size'],
            shuffle=True,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False
        ),
        'val': torch.utils.data.DataLoader(
            datasets['val'],
            batch_size=cfg['data']['batch_size'],
            shuffle=False,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False
        ),
        'test': torch.utils.data.DataLoader(
            datasets['test'],
            batch_size=1,
            shuffle=False,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False
        )
    }


    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name';   sheet['A2'] = exp_name
        sheet['B1'] = 'batch size'; sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs';     sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim';    sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)

    # ---------- Model ----------
    G = CoupledGenerator(
        in_ch=1,
        base_ch=cfg['model'].get('dim', 64),
        n_down=cfg['model'].get('n_downsample', 2),
        n_res=cfg['model'].get('n_residual', 4),
        out_activation=cfg['model'].get('out_activation', 'sigmoid')
    ).to(device)

    D = CoupledDiscriminator(in_ch=1).to(device)

    G.apply(weights_init_normal)
    D.apply(w_init_D)

    lr = cfg['trainer']['optimizer']['lr']
    opt_G = optim.Adam(G.parameters(), lr=lr, betas=(0.5, 0.999))
    opt_D = optim.Adam(D.parameters(), lr=lr, betas=(0.5, 0.999))

    l1      = nn.L1Loss().to(device)
    mse_gan = nn.MSELoss().to(device)
    early_stopping_criterion = nn.L1Loss().to(device)

    # ---------- Checkpoint paths ----------
    ckpt_G_path = os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_G'])
    ckpt_D_path = os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_D'])
    has_ckpt = os.path.exists(ckpt_G_path) and os.path.exists(ckpt_D_path)

    history = None
    resume_state = None
    training_finished = False

    # ============================================================
    # 3 cases:
    # 1) No checkpoint -> training from scratch
    # 2) Checkpoint present but training NOT finished -> resume
    # 3) Checkpoint present and training FINISHED -> testing only
    # ============================================================
    if has_ckpt:
        ckpt_G = util_model.load_cogan_checkpoint_G(ckpt_G_path, device)
        training_finished = ckpt_G.get("training_finished", False)

        if training_finished:

            print(f"[Fold {fold}] Checkpoint trovato e training già terminato: eseguo SOLO il test.")
            G.load_state_dict(ckpt_G["state_dict"])
            ckpt_D = torch.load(ckpt_D_path, map_location=device)
            D.load_state_dict(ckpt_D["state_dict"])
        else:

            print(f"[Fold {fold}] Checkpoint trovato ma training NON finito: riprendo da dove si era fermato.")
            G.load_state_dict(ckpt_G["state_dict"])
            opt_G.load_state_dict(ckpt_G["optimizer"])
            for pg in opt_G.param_groups:
                pg["lr"] = lr

            ckpt_D = torch.load(ckpt_D_path, map_location=device)
            D.load_state_dict(ckpt_D["state_dict"])
            opt_D.load_state_dict(ckpt_D["optimizer"])
            for pg in opt_D.param_groups:
                pg["lr"] = lr

            start_epoch       = ckpt_G.get("epoch", -1) + 1
            best_loss         = ckpt_G.get("best_loss", float("inf"))
            best_epoch        = ckpt_G.get("best_epoch", -1)
            epochs_no_improve = ckpt_G.get("epochs_no_improve", 0)

            resume_state = {
                "start_epoch": start_epoch,
                "best_loss": best_loss,
                "best_epoch": best_epoch,
                "epochs_no_improve": epochs_no_improve,
            }
    else:

        print(f"[Fold {fold}] Nessun checkpoint: inizio training da zero.")

    # ---------- TRAIN  ----------
    if (not has_ckpt) or (has_ckpt and not training_finished):
        G, D, history = util_model.train_cogan(
            G, D,
            data_loaders=data_loaders,
            early_stopping_criterion=early_stopping_criterion,
            opt_G=opt_G, opt_D=opt_D,
            l1=l1, mse_gan=mse_gan,
            model_fold_dir=model_fold_dir,
            cfg_trainer=cfg['trainer'],
            device=device,
            resume_state=resume_state  # supporto resume
        )

        if history is not None:
            util_model.plot_training(history, model_name, plot_training_fold_dir)
    else:
        print(f"[Fold {fold}] Salto training (già concluso in run precedente).")

    # ==========================
    # (test_cogan)
    # ==========================
    mse_metric  = nn.MSELoss().to(device)
    psnr_metric = PeakSignalNoiseRatio().to(device)

    results = util_model.test_cogan(
        G,
        loader_test=data_loaders['test'],
        mse_metric=mse_metric,
        psnr_metric=psnr_metric,
        device=device,
        outputs_dir=outputs_fold_dir,
        save_outputs=True,
        amp=True,
        lpips_backbone="alex",
        is_splits=10,
        return_per_image=False
    )

    print(
        f"Fold {fold} → "
        f"MSE={results['mse']:.6f}, PSNR={results['psnr']:.4f}, "
        f"VIF={results['vif']:.4f}, SSIM={results['ssim']:.4f}, "
        f"LPIPS={results['lpips']:.4f}, FID={results['fid']:.4f}, "
        f"IS={results['is_mean']:.4f}±{results['is_std']:.4f}"
    )


    fold_metrics_full.append({
        "fold": fold,
        "mse": results["mse"],
        "psnr": results["psnr"],
        "vif": results["vif"],
        "ssim": results["ssim"],
        "lpips": results["lpips"],
        "fid": results["fid"],
        "is_mean": results["is_mean"],
        "is_std": results["is_std"],
    })


if fold_env is None and len(fold_metrics_full) == cv:
    df_folds = pd.DataFrame(fold_metrics_full).set_index("fold").sort_index()

    mse_m,   mse_std   = df_folds["mse"].mean(),      df_folds["mse"].std(ddof=1)
    psnr_m,  psnr_std  = df_folds["psnr"].mean(),     df_folds["psnr"].std(ddof=1)
    vif_m,   vif_std   = df_folds["vif"].mean(),      df_folds["vif"].std(ddof=1)
    ssim_m,  ssim_std  = df_folds["ssim"].mean(),     df_folds["ssim"].std(ddof=1)
    lpips_m, lpips_std = df_folds["lpips"].mean(),    df_folds["lpips"].std(ddof=1)
    fid_m,   fid_std   = df_folds["fid"].mean(),      df_folds["fid"].std(ddof=1)
    is_m,    is_std    = df_folds["is_mean"].mean(),  df_folds["is_mean"].std(ddof=1)

    def fcol(metric, i):
        return float(df_folds.loc[i, metric])

    row_out = {
        "MSE_m": mse_m, "MSE_std": mse_std,
        "PSNR_m": psnr_m, "PSNR_std": psnr_std,
        "VIF_m": vif_m, "VIF_std": vif_std,
        "SSIM_m": ssim_m, "SSIM_std": ssim_std,
        "LPIPS_m": lpips_m, "LPIPS_std": lpips_std,
        "FID_m": fid_m, "FID_std": fid_std,
        "IS_m": is_m, "IS_std": is_std,
    }

    for i in range(cv): row_out[f"MSE_{i}"]   = fcol("mse", i)
    for i in range(cv): row_out[f"PSNR_{i}"]  = fcol("psnr", i)
    for i in range(cv): row_out[f"VIF_{i}"]   = fcol("vif", i)
    for i in range(cv): row_out[f"SSIM_{i}"]  = fcol("ssim", i)
    for i in range(cv): row_out[f"LPIPS_{i}"] = fcol("lpips", i)
    for i in range(cv): row_out[f"FID_{i}"]   = fcol("fid", i)
    for i in range(cv): row_out[f"IS_{i}"]    = fcol("is_mean", i)

    ordered_cols = [
        "MSE_m","MSE_std","PSNR_m","PSNR_std","VIF_m","VIF_std","SSIM_m","SSIM_std",
        "LPIPS_m","LPIPS_std","FID_m","FID_std","IS_m","IS_std",
        "MSE_0","MSE_1","MSE_2","MSE_3","MSE_4",
        "PSNR_0","PSNR_1","PSNR_2","PSNR_3","PSNR_4",
        "VIF_0","VIF_1","VIF_2","VIF_3","VIF_4",
        "SSIM_0","SSIM_1","SSIM_2","SSIM_3","SSIM_4",
        "LPIPS_0","LPIPS_1","LPIPS_2","LPIPS_3","LPIPS_4",
        "FID_0","FID_1","FID_2","FID_3","FID_4",
        "IS_0","IS_1","IS_2","IS_3","IS_4",
    ]

    df_out = pd.DataFrame([[row_out[c] for c in ordered_cols]], columns=ordered_cols)
    excel_path = os.path.join(report_dir, f"{exp_name_excel}_test.xlsx")
    df_out.to_excel(excel_path, index=False)
    print("\n[OK] Excel (1 riga) salvato in:", excel_path)
else:
    print("\n[INFO] Excel 1-riga globale NON generato (o stai usando FOLD_IDX, o non hai tutte le fold in questo run).")
