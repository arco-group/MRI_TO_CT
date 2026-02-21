import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])
import os, yaml, ssl, collections
import torch
import torch.nn as nn
import torch.optim as optim
import pandas as pd
from sys import platform
from openpyxl import Workbook, load_workbook
from src.model.Cogan.generator_model import CoupledGenerator, weights_init_normal
from src.model.Cogan.discriminator_model import CoupledDiscriminator, weights_init_normal as w_init_D
import model.Cogan.codici_vecchi.util_model_vecchio as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# Config
if platform == 'win32':
    cfg_path = "./configs/cogan_train.yaml"
else:
    cfg_path = "/mimer/NOBACKUP/groups/snic2022-5-277/rofena/Progetto_Alessandro/configs/cogan_train.yaml"

with open(cfg_path) as f:
    cfg = yaml.load(f, Loader=yaml.FullLoader)

exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
cv = cfg['data']['cv']
fold_list = list(range(cv))

device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type=="cpu" else cfg['device']['gpu_num_workers']
print("Device:", device)

fold_dir = os.path.join(cfg['data']['fold_dir'])

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

plot_training_dir = os.path.join(report_dir, "training")
outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(plot_training_dir)
util_general.create_dir(outputs_dir)

report_file_mse  = os.path.join(report_dir, 'report_mse.xlsx')
report_file_psnr = os.path.join(report_dir, 'report_psnr.xlsx')
report_file_vif  = os.path.join(report_dir, 'report_vif.xlsx')
report_file_ssim = os.path.join(report_dir, 'report_ssim.xlsx')

results_mse  = collections.defaultdict(lambda: [])
results_psnr = collections.defaultdict(lambda: [])
results_vif  = collections.defaultdict(lambda: [])
results_ssim = collections.defaultdict(lambda: [])
metric_cols_mse = []
metric_cols_psnr=[]
metric_cols_vif=[]
metric_cols_ssim=[]

for fold in fold_list:
    print(f"\n========== Fold {fold} ==========")
    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)

    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_G'])):
        print(f"Skipping Fold {fold} (checkpoint found)")
        continue

    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    util_general.create_dir(outputs_fold_dir)

    metric_cols_mse.append(f"{fold} MSE")
    metric_cols_psnr.append(f"{fold} PSNR")
    metric_cols_vif.append(f"{fold} VIF")
    metric_cols_ssim.append(f"{fold} SSIM")

    # Data loaders ALE (k-fold CSV)
    fold_data = {step: pd.read_csv(os.path.join(fold_dir, str(fold), f'{step}.csv'), index_col='id_slice') for step in ['train','val','test']}
    datasets = {step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train','val','test']}
    data_loaders = {
        'train': torch.utils.data.DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'],
                                             shuffle=True,
                                             num_workers=num_workers,
                                             worker_init_fn=util_data.seed_worker,
                                             pin_memory=True,
                                             persistent_workers=True if num_workers>0 else False),
        'val': torch.utils.data.DataLoader(datasets['val'],
                                           batch_size=cfg['data']['batch_size'],
                                           shuffle=False,
                                           num_workers=num_workers,
                                           worker_init_fn=util_data.seed_worker,
                                           pin_memory=True,
                                           persistent_workers=True if num_workers>0 else False),
        'test': torch.utils.data.DataLoader(datasets['test'],
                                            batch_size=1,
                                            shuffle=False,
                                            num_workers=num_workers,
                                            worker_init_fn=util_data.seed_worker,
                                            pin_memory=True,
                                            persistent_workers=True if num_workers>0 else False)
    }

    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name'
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size'
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs'
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim'
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)

    # Models
    G = CoupledGenerator(
        in_ch=1,
        base_ch=cfg['model'].get('dim',64),
        n_down=cfg['model'].get('n_downsample',2),
        n_res=cfg['model'].get('n_residual',4),
        out_activation=cfg['model'].get('out_activation','sigmoid')
    ).to(device)

    D = CoupledDiscriminator(in_ch=1).to(device)

    G.apply(weights_init_normal)
    D.apply(w_init_D)

    lr = cfg['trainer']['optimizer']['lr']
    opt_G = optim.Adam(G.parameters(), lr=lr, betas=(0.5,0.999))
    opt_D = optim.Adam(D.parameters(), lr=lr, betas=(0.5,0.999))

    l1 = nn.L1Loss().to(device)
    mse_gan = nn.MSELoss().to(device)
    early_stopping_criterion = nn.L1Loss().to(device)

    # Train
    G, D, history = util_model.train_cogan(
        G, D,
        data_loaders=data_loaders,
        early_stopping_criterion=early_stopping_criterion,
        opt_G=opt_G, opt_D=opt_D,
        l1=l1, mse_gan=mse_gan,
        model_fold_dir=model_fold_dir,
        cfg_trainer=cfg['trainer'],
        device=device
    )

    util_model.plot_training(history, model_name, plot_training_fold_dir)

    # Test MRI->CT
    mse_t, psnr_t, vif_t, ssim_t = util_model.test_4metrics_cogan(
        G, loader_test=data_loaders['test'], device=device,
        outputs_dir=outputs_fold_dir, save_output=True
    )
    print(f"Fold {fold} - MSE: {mse_t:.6f}, PSNR: {psnr_t:.4f}, VIF: {vif_t:.4f}, SSIM: {ssim_t:.4f}")

    results_mse[f"{fold} MSE"].append(mse_t)
    results_psnr[f"{fold} PSNR"].append(psnr_t)
    results_vif[f"{fold} VIF"].append(vif_t)
    results_ssim[f"{fold} SSIM"].append(ssim_t)

    # Save reports
    def _save_report(results_dict, metric_cols, fname):
        frame = pd.DataFrame(results_dict)
        frame.insert(0,'std', frame[metric_cols].std(axis=1))
        frame.insert(0,'mean', frame[metric_cols].mean(axis=1))
        frame.insert(0,'model', model_name)
        frame.to_excel(fname, index=False)

    _save_report(results_mse, metric_cols_mse, report_file_mse)
    _save_report(results_psnr, metric_cols_psnr, report_file_psnr)
    _save_report(results_vif, metric_cols_vif, report_file_vif)
    _save_report(results_ssim, metric_cols_ssim, report_file_ssim)
