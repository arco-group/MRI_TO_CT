import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])  # permette import relativi al progetto

import os
import yaml
import ssl
import torch
import torch.nn as nn
import torch.optim as optim
import pandas as pd
import collections
from torchmetrics import PeakSignalNoiseRatio
from sys import platform

from src.model.Unit.generator_model import Encoder, Generator, ResidualBlock
from src.model.Unit.discriminator_model import Discriminator
import src.model.Unit.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data
from openpyxl import Workbook
from openpyxl import load_workbook

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

#lettura e configurazione
if platform == 'win32':
    cfg_path = "./configs/unit_train.yaml"
else:
    cfg_path = "/mimer/NOBACKUP/groups/snic2022-5-277/rofena/Progetto_Alessandro/configs/unit_train.yaml"

with open(cfg_path) as f:
    cfg = yaml.load(f, Loader=yaml.FullLoader)

exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
cv = cfg['data']['cv']
fold_list = list(range(cv))
learning_rate = cfg['trainer']['optimizer']['lr']

# Se esiste la variabile d'ambiente FOLD_IDX, uso solo quella fold
fold_env = os.environ.get("FOLD_IDX", None)
if fold_env is not None:
    fold_list = [int(fold_env)]
    print(f"*** Esecuzione solo per la fold {fold_list[0]} (da variabile d'ambiente FOLD_IDX) ***")
else:
    cv = cfg['data']['cv']
    fold_list = list(range(cv))
    print(f"*** Esecuzione per tutte le fold: {fold_list} ***")

device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(f"Device: {device}")

# Files and Directories
fold_dir = os.path.join(cfg['data']['fold_dir'])

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

report_file_mse = os.path.join(report_dir, 'report_mse.xlsx')
report_file_psnr = os.path.join(report_dir, 'report_psnr.xlsx')
report_file_vif = os.path.join(report_dir, 'report_vif.xlsx')
report_file_ssim = os.path.join(report_dir, 'report_ssim.xlsx')

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

# CV
results_mse = collections.defaultdict(lambda: [])
metric_cols_mse = []
results_psnr = collections.defaultdict(lambda: [])
metric_cols_psnr = []
results_vif = collections.defaultdict(lambda: [])
metric_cols_vif = []
results_ssim = collections.defaultdict(lambda: [])
metric_cols_ssim = []

#training
for fold in fold_list:
    print(f"\n===== FOLD {fold} =====")

    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)

    # Non skippare: affidiamo la logica di resume a train_unit()
    # Se vuoi, puoi solo stampare se esiste già un qualche checkpoint
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_D2'])):
        print(f"[Fold {fold}] Trovato un checkpoint D2: il training potrà riprendere dai pesi salvati.")

    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    util_general.create_dir(outputs_fold_dir)

    metric_cols_mse.append(f"{fold} MSE")
    metric_cols_psnr.append(f"{fold} PSNR")
    metric_cols_vif.append(f"{fold} VIF")
    metric_cols_ssim.append(f"{fold} SSIM")

    pf = 4 if num_workers > 0 else None
    #Datalodears
    fold_data = {step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col="id_slice") for step in ['train', 'val', 'test']}
    datasets = {step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}
    data_loaders = {
        'train': torch.utils.data.DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker,pin_memory=True,
        persistent_workers=True if num_workers > 0 else False, prefetch_factor=pf if num_workers > 0 else None,),
        'val': torch.utils.data.DataLoader(datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker,pin_memory=True,
        persistent_workers=True if num_workers > 0 else False, prefetch_factor=pf if num_workers > 0 else None,),
        'test': torch.utils.data.DataLoader(datasets['test'], batch_size=1, shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker,pin_memory=True,
        persistent_workers=True if num_workers > 0 else False, prefetch_factor=pf if num_workers > 0 else None,)
    }

    # Crea info.xlsx solo al primo fold
    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name';
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size';
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs';
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim';
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)

    #inizializzazione modelli
    shared_dim = cfg['model']['dim'] * (2 ** cfg['model']['n_downsample'])
    shared_E = ResidualBlock(features=shared_dim).to(device)
    shared_G = ResidualBlock(features=shared_dim).to(device)

    E1 = Encoder(in_channels=1, dim=cfg['model']['dim'], n_downsample=cfg['model']['n_downsample'], shared_block=shared_E).to(device)
    E2 = Encoder(in_channels=1, dim=cfg['model']['dim'], n_downsample=cfg['model']['n_downsample'], shared_block=shared_E).to(device)
    G1 = Generator(out_channels=1, dim=cfg['model']['dim'], n_upsample=cfg['model']['n_downsample'],
                   shared_block=shared_G, out_activation='sigmoid').to(device)
    G2 = Generator(out_channels=1, dim=cfg['model']['dim'], n_upsample=cfg['model']['n_downsample'],
                   shared_block=shared_G, out_activation='sigmoid').to(device)
    D1 = Discriminator(in_channels=1).to(device)
    D2 = Discriminator(in_channels=1).to(device)

    # Ottimizzatori
    lr = cfg['trainer']['optimizer']['lr']
    opt_G = optim.Adam(list(E1.parameters()) + list(E2.parameters()) + list(G1.parameters()) + list(G2.parameters()), lr=lr, betas=(0.5, 0.999))
    opt_D1 = optim.Adam(D1.parameters(), lr=lr, betas=(0.5, 0.999))
    opt_D2 = optim.Adam(D2.parameters(), lr=lr, betas=(0.5, 0.999))

    mse, l1 = nn.MSELoss().to(device), nn.L1Loss().to(device)
    early_stopping_criterion = nn.L1Loss().to(device)

#training
    E1, E2, G1, G2, D1, D2, history = util_model.train_unit(
        E1, E2, G1, G2, D1, D2,
        data_loaders=data_loaders, early_stopping_criterion=early_stopping_criterion,
        opt_G=opt_G, opt_D1=opt_D1, opt_D2=opt_D2,
        l1_pix=l1, mse_gan=mse,
        model_fold_dir=model_fold_dir, cfg_trainer=cfg['trainer'], device=device
    )

    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    psnr = PeakSignalNoiseRatio().to(device)
    mse_t, psnr_t, vif_t, ssim_t = util_model.test_4metrics_unit(E1, G2, loader=data_loaders['test'], mse=mse, psnr=psnr, device=device, outputs_dir=outputs_fold_dir, save_output=True)
    print(f"FOLD {fold} → MSE: {mse_t:.4f}, PSNR: {psnr_t:.4f}, VIF: {vif_t:.4f}, SSIM: {ssim_t:.4f}")

    # Update report
    results_mse["%s MSE" % str(fold)].append(mse_t)
    results_psnr["%s PSNR" % str(fold)].append(psnr_t)
    results_vif["%s VIF" % str(fold)].append(vif_t)
    results_ssim["%s SSIM" % str(fold)].append(ssim_t)

    # Save Results MSE
    results_mse_frame = pd.DataFrame(results_mse)
    results_mse_frame.insert(loc=0, column='std MSE', value=results_mse_frame[metric_cols_mse].std(axis=1))
    results_mse_frame.insert(loc=0, column='mean MSE', value=results_mse_frame[metric_cols_mse].mean(axis=1))
    results_mse_frame.insert(loc=0, column='model', value=model_name)
    results_mse_frame.to_excel(report_file_mse, index=False)

    # Save Results PSNR
    results_psnr_frame = pd.DataFrame(results_psnr)
    results_psnr_frame.insert(loc=0, column='std PSNR', value=results_psnr_frame[metric_cols_psnr].std(axis=1))
    results_psnr_frame.insert(loc=0, column='mean PSNR', value=results_psnr_frame[metric_cols_psnr].mean(axis=1))
    results_psnr_frame.insert(loc=0, column='model', value=model_name)
    results_psnr_frame.to_excel(report_file_psnr, index=False)

    # Save Results VIF
    results_vif_frame = pd.DataFrame(results_vif)
    results_vif_frame.insert(loc=0, column='std VIF', value=results_vif_frame[metric_cols_vif].std(axis=1))
    results_vif_frame.insert(loc=0, column='mean VIF', value=results_vif_frame[metric_cols_vif].mean(axis=1))
    results_vif_frame.insert(loc=0, column='model', value=model_name)
    results_vif_frame.to_excel(report_file_vif, index=False)

    # Save Results SSIM
    results_ssim_frame = pd.DataFrame(results_ssim)
    results_ssim_frame.insert(loc=0, column='std SSIM', value=results_ssim_frame[metric_cols_ssim].std(axis=1))
    results_ssim_frame.insert(loc=0, column='mean SSIM', value=results_ssim_frame[metric_cols_ssim].mean(axis=1))
    results_ssim_frame.insert(loc=0, column='model', value=model_name)
    results_ssim_frame.to_excel(report_file_ssim, index=False)


