import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])

import os
import yaml
import ssl
import torch
import torch.nn as nn
import torch.optim as optim
import pandas as pd
import collections
from torchmetrics import PeakSignalNoiseRatio
from sys import platform
import numpy as np
from src.model.Unit.generator_model import Encoder, Generator, ResidualBlock
from src.model.Unit.discriminator_model import Discriminator
import src.model.Unit.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data
from openpyxl import Workbook, load_workbook

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context


if platform == 'win32':
    cfg_path = "./configs/unit_train.yaml"
else:
    cfg_path = ".../configs/unit_train.yaml"

with open(cfg_path) as f:
    cfg = yaml.load(f, Loader=yaml.FullLoader)

exp_name       = cfg['exp_name']
exp_name_excel = cfg['exp_name_excel']
print(exp_name)
model_name = cfg['model']['model_name']

cv = cfg['data']['cv']
fold_list = list(range(cv))
learning_rate = cfg['trainer']['optimizer']['lr']


fold_env = os.environ.get("FOLD_IDX", None)
if fold_env is not None:
    fold_list = [int(fold_env)]
    print(f"*** Execution on the Fold {fold_list[0]} (FOLD_IDX) ***")
else:
    cv = cfg['data']['cv']
    fold_list = list(range(cv))
    print(f"*** Execution on all folds: {fold_list} ***")

device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(f"Device: {device}")


fold_dir = os.path.join(cfg['data']['fold_dir'])

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

save_img_output = cfg.get('output', {}).get('save_img_output', True)


fold_metrics = {
    "mse":   [],
    "psnr":  [],
    "vif":   [],
    "ssim":  [],
    "lpips": [],
    "fid":   [],
    "is":    []  # IS_mean
}


for fold in fold_list:
    print(f"\n===== FOLD {fold} =====")

    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)

    # Solo info: controllo se c'è già un checkpoint D2 (significa che è stato salvato qualcosa)
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_D2'])):
        print(f"[Fold {fold}] Trovato un checkpoint D2: il training potrà riprendere dai pesi salvati (via MAIN checkpoint).")

    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    outputs_fold_dir       = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    util_general.create_dir(outputs_fold_dir)

    # -----------------------------------------------------------------
    # DATASET / DATALOADER
    # -----------------------------------------------------------------
    pf = 4 if num_workers > 0 else None

    fold_data = {
        step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col="id_slice")
        for step in ['train', 'val', 'test']
    }

    datasets = {
        step: util_data.ImgDataset(
            data=fold_data[step],
            cfg_data=cfg['data'],
            step=step,
            do_augmentation=True
        )
        for step in ['train', 'val', 'test']
    }

    data_loaders = {
        'train': torch.utils.data.DataLoader(
            datasets['train'],
            batch_size=cfg['data']['batch_size'],
            shuffle=True,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False,
            prefetch_factor=pf if num_workers > 0 else None,
        ),
        'val': torch.utils.data.DataLoader(
            datasets['val'],
            batch_size=cfg['data']['batch_size'],
            shuffle=False,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False,
            prefetch_factor=pf if num_workers > 0 else None,
        ),
        'test': torch.utils.data.DataLoader(
            datasets['test'],
            batch_size=1,
            shuffle=False,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False,
            prefetch_factor=pf if num_workers > 0 else None,
        ),
    }


    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name';    sheet['A2'] = exp_name
        sheet['B1'] = 'batch size';  sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs';      sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim';     sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)


    shared_dim = cfg['model']['dim'] * (2 ** cfg['model']['n_downsample'])
    shared_E   = ResidualBlock(features=shared_dim).to(device)
    shared_G   = ResidualBlock(features=shared_dim).to(device)

    E1 = Encoder(
        in_channels=1,
        dim=cfg['model']['dim'],
        n_downsample=cfg['model']['n_downsample'],
        shared_block=shared_E
    ).to(device)

    E2 = Encoder(
        in_channels=1,
        dim=cfg['model']['dim'],
        n_downsample=cfg['model']['n_downsample'],
        shared_block=shared_E
    ).to(device)

    G1 = Generator(
        out_channels=1,
        dim=cfg['model']['dim'],
        n_upsample=cfg['model']['n_downsample'],
        shared_block=shared_G,
        out_activation='sigmoid'
    ).to(device)

    G2 = Generator(
        out_channels=1,
        dim=cfg['model']['dim'],
        n_upsample=cfg['model']['n_downsample'],
        shared_block=shared_G,
        out_activation='sigmoid'
    ).to(device)

    D1 = Discriminator(in_channels=1).to(device)
    D2 = Discriminator(in_channels=1).to(device)


    lr = cfg['trainer']['optimizer']['lr']
    opt_G  = optim.Adam(
        list(E1.parameters()) + list(E2.parameters()) + list(G1.parameters()) + list(G2.parameters()),
        lr=lr, betas=(0.5, 0.999)
    )
    opt_D1 = optim.Adam(D1.parameters(), lr=lr, betas=(0.5, 0.999))
    opt_D2 = optim.Adam(D2.parameters(), lr=lr, betas=(0.5, 0.999))

    mse = nn.MSELoss().to(device)
    l1  = nn.L1Loss().to(device)
    early_stopping_criterion = nn.L1Loss().to(device)


    E1, E2, G1, G2, D1, D2, history = util_model.train_unit(
        E1, E2, G1, G2, D1, D2,
        data_loaders=data_loaders,
        early_stopping_criterion=early_stopping_criterion,
        opt_G=opt_G, opt_D1=opt_D1, opt_D2=opt_D2,
        l1_pix=l1, mse_gan=mse,
        model_fold_dir=model_fold_dir,
        cfg_trainer=cfg['trainer'],
        device=device
    )


    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)


    pretrained_model_dir = model_fold_dir

    def _safe_load(ckpt_name, model, optimizer, lr, device):
        ckpt_path = os.path.join(pretrained_model_dir, ckpt_name)
        if os.path.exists(ckpt_path):
            print(f"=> Loading checkpoint: {ckpt_path}")
            util_general.load_checkpoint(ckpt_path, model, optimizer, lr, device)
        else:
            print(f"[ATTENZIONE] Checkpoint {ckpt_path} NON trovato. "
                  f"Uso i pesi correnti in memoria.")


    _safe_load(cfg['trainer']['CHECKPOINT_E1'], E1, opt_G, lr, device)
    _safe_load(cfg['trainer']['CHECKPOINT_E2'], E2, opt_G, lr, device)
    _safe_load(cfg['trainer']['CHECKPOINT_G1'], G1, opt_G, lr, device)
    _safe_load(cfg['trainer']['CHECKPOINT_G2'], G2, opt_G, lr, device)


    mse_metric  = nn.MSELoss().to(device)
    psnr_metric = PeakSignalNoiseRatio().to(device)


    results = util_model.test_unit(
        E1,          # encoder MRI
        G2,          # generator CT
        data_loaders['test'],  # loader_test
        mse_metric,
        psnr_metric,
        device,
        outputs_dir=outputs_fold_dir,
        save_outputs=save_img_output,
        return_per_image=True
    )

    fold_metrics["mse"].append(results["mse"])
    fold_metrics["psnr"].append(results["psnr"])
    fold_metrics["vif"].append(results["vif"])
    fold_metrics["ssim"].append(results["ssim"])
    fold_metrics["lpips"].append(results["lpips"])
    fold_metrics["fid"].append(results["fid"])
    fold_metrics["is"].append(results["is_mean"])

    print(f"[fold {fold}] mse={results['mse']:.6f} psnr={results['psnr']:.4f} "
          f"vif={results['vif']:.4f} ssim={results['ssim']:.4f} "
          f"lpips={results['lpips']:.4f} fid={results['fid']:.4f} is_mean={results['is_mean']:.4f}")



def _mean_std(x):
    s = pd.Series(x, dtype="float64")
    return float(s.mean()), float(s.std(ddof=1))

MSE_m,   MSE_std   = _mean_std(fold_metrics["mse"])
PSNR_m,  PSNR_std  = _mean_std(fold_metrics["psnr"])
VIF_m,   VIF_std   = _mean_std(fold_metrics["vif"])
SSIM_m,  SSIM_std  = _mean_std(fold_metrics["ssim"])
LPIPS_m, LPIPS_std = _mean_std(fold_metrics["lpips"])
FID_m,   FID_std   = _mean_std(fold_metrics["fid"])
IS_m,    IS_std    = _mean_std(fold_metrics["is"])

def _pad_to_cv(vals, cv, fill=np.nan):
    vals = list(vals)
    if len(vals) < cv:
        vals = vals + [fill] * (cv - len(vals))
    return vals[:cv]

cv = cfg['data']['cv']  # usa quello del config

mse_list   = _pad_to_cv(fold_metrics["mse"],   cv)
psnr_list  = _pad_to_cv(fold_metrics["psnr"],  cv)
vif_list   = _pad_to_cv(fold_metrics["vif"],   cv)
ssim_list  = _pad_to_cv(fold_metrics["ssim"],  cv)
lpips_list = _pad_to_cv(fold_metrics["lpips"], cv)
fid_list   = _pad_to_cv(fold_metrics["fid"],   cv)
is_list    = _pad_to_cv(fold_metrics["is"],    cv)

# ora puoi creare MSE_0..MSE_4 in modo safe
MSE_0, MSE_1, MSE_2, MSE_3, MSE_4 = mse_list
PSNR_0,PSNR_1,PSNR_2,PSNR_3,PSNR_4 = psnr_list
VIF_0, VIF_1, VIF_2, VIF_3, VIF_4 = vif_list
SSIM_0,SSIM_1,SSIM_2,SSIM_3,SSIM_4 = ssim_list
LPIPS_0,LPIPS_1,LPIPS_2,LPIPS_3,LPIPS_4 = lpips_list
FID_0,FID_1,FID_2,FID_3,FID_4 = fid_list
IS_0,IS_1,IS_2,IS_3,IS_4 = is_list

ordered_row = [
    MSE_m, MSE_std,
    PSNR_m, PSNR_std,
    VIF_m, VIF_std,
    SSIM_m, SSIM_std,
    LPIPS_m, LPIPS_std,
    FID_m, FID_std,
    IS_m, IS_std,

    MSE_0, MSE_1, MSE_2, MSE_3, MSE_4,
    PSNR_0, PSNR_1, PSNR_2, PSNR_3, PSNR_4,
    VIF_0, VIF_1, VIF_2, VIF_3, VIF_4,
    SSIM_0, SSIM_1, SSIM_2, SSIM_3, SSIM_4,
    LPIPS_0, LPIPS_1, LPIPS_2, LPIPS_3, LPIPS_4,
    FID_0, FID_1, FID_2, FID_3, FID_4,
    IS_0, IS_1, IS_2, IS_3, IS_4
]

ordered_cols = [
    "MSE_m","MSE_std",
    "PSNR_m","PSNR_std",
    "VIF_m","VIF_std",
    "SSIM_m","SSIM_std",
    "LPIPS_m","LPIPS_std",
    "FID_m","FID_std",
    "IS_m","IS_std",

    "MSE_0","MSE_1","MSE_2","MSE_3","MSE_4",
    "PSNR_0","PSNR_1","PSNR_2","PSNR_3","PSNR_4",
    "VIF_0","VIF_1","VIF_2","VIF_3","VIF_4",
    "SSIM_0","SSIM_1","SSIM_2","SSIM_3","SSIM_4",
    "LPIPS_0","LPIPS_1","LPIPS_2","LPIPS_3","LPIPS_4",
    "FID_0","FID_1","FID_2","FID_3","FID_4",
    "IS_0","IS_1","IS_2","IS_3","IS_4"
]

df_out = pd.DataFrame([ordered_row], columns=ordered_cols)

excel_path = os.path.join(report_dir, f"{exp_name_excel}_test_summary.xlsx")
with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    df_out.to_excel(writer, sheet_name="metrics", index=False)

print(f"\n[OK] Excel salvato (1 riga ordinata) in: {excel_path}")
