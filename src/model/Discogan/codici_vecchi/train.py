import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])  # permette import relativi al progetto
import torch
from torch.utils.data import DataLoader
import torch.nn as nn
import torch.optim as optim
import itertools
import os
import yaml
import ssl
import pandas as pd
import collections
from torchmetrics import PeakSignalNoiseRatio
from sys import platform
from models import *
from src.model.Discogan.models import GeneratorUNet, Discriminator
import src.model.Discogan.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data
from openpyxl import Workbook
from openpyxl import load_workbook


torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

#Configuration File
if platform == 'win32':
    cfg_path = "./configs/discogan_train.yaml"
else:
    cfg_path = "/mimer/NOBACKUP/groups/snic2022-5-277/rofena/Progetto_Alessandro/configs/discogan_train.yaml"

with open(cfg_path) as f:
    cfg = yaml.load(f, Loader=yaml.FullLoader)

#Parameters
exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
lr = cfg['trainer']['optimizer']['lr']
cv = cfg['data']['cv']
fold_list = list(range(cv))

#device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(f"Device: {device}")

#Files and Directories
fold_dir = cfg['data']['fold_dir']

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
plot_training_dir = os.path.join(report_dir, "training")
outputs_dir = os.path.join(report_dir, "outputs")

for d in [model_dir, report_dir, plot_training_dir, outputs_dir]:
    util_general.create_dir(d)

report_file_mse = os.path.join(report_dir, 'report_mse.xlsx')
report_file_psnr = os.path.join(report_dir, 'report_psnr.xlsx')
report_file_vif = os.path.join(report_dir, 'report_vif.xlsx')
report_file_ssim = os.path.join(report_dir, 'report_ssim.xlsx')

#CV
results_mse = collections.defaultdict(lambda: [])
metric_cols_mse = []
results_psnr = collections.defaultdict(lambda: [])
metric_cols_psnr = []
results_vif = collections.defaultdict(lambda: [])
metric_cols_vif = []
results_ssim = collections.defaultdict(lambda: [])
metric_cols_ssim = []

#folds iteration
for fold in fold_list:
    print(f"\n===== FOLD {fold} =====")

    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_DISC_B'])):
        print("Skipping Fold %d" % fold)
        continue
    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    util_general.create_dir(outputs_fold_dir)

    #Results Frame
    metric_cols_mse.append(f"{fold} MSE")
    metric_cols_psnr.append(f"{fold} PSNR")
    metric_cols_vif.append(f"{fold} VIF")
    metric_cols_ssim.append(f"{fold} SSIM")

   #Data Loaders
    fold_data = {step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col="id_slice") for step in ['train', 'val', 'test']}
    datasets = {step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}
    data_loaders = {
        'train': torch.utils.data.DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'val': torch.utils.data.DataLoader(datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'test': torch.utils.data.DataLoader(datasets['test'], batch_size=1, shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker)}

    # Creo file excel con info generali
    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        # Definisco le colonne del file excel
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name'
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size'
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs'
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim'
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)


    #inizializzazione modelli
    #input_shape = (opt.channels, opt.img_height, opt.img_width)
    input_shape = (1, cfg['data']['img_dim'], cfg['data']['img_dim'])

    # Initialize generator and discriminator
    adversarial_loss = torch.nn.MSELoss().to(device)
    cycle_loss = torch.nn.L1Loss().to(device)
    pixelwise_loss = torch.nn.L1Loss().to(device)

    G_AB = GeneratorUNet(input_shape).to(device)
    G_BA = GeneratorUNet(input_shape).to(device)
    D_A = Discriminator(input_shape).to(device)
    D_B = Discriminator(input_shape).to(device)

    G_AB.apply(weights_init_normal)
    G_BA.apply(weights_init_normal)
    D_A.apply(weights_init_normal)
    D_B.apply(weights_init_normal)

    # Ottimizzatori

    optimizer_G = torch.optim.Adam(itertools.chain(G_AB.parameters(), G_BA.parameters()), lr=lr, betas=(0.5, 0.999))
    optimizer_D_A = torch.optim.Adam(D_A.parameters(), lr=lr, betas=(0.5, 0.999))
    optimizer_D_B = torch.optim.Adam(D_B.parameters(), lr=lr, betas=(0.5, 0.999))

    # Loss
    adversarial_loss = torch.nn.MSELoss().to(device)
    cycle_loss = torch.nn.L1Loss().to(device)
    pixelwise_loss = torch.nn.L1Loss().to(device)

    early_stopping_criterion = nn.L1Loss().to(device)

   #Train model
    G_AB, G_BA, D_A, D_B, history = util_model.train_discogan(
        G_AB, G_BA, D_A, D_B,
        data_loaders=data_loaders, early_stopping_criterion=early_stopping_criterion,
        optimizer_G=optimizer_G, optimizer_D_A=optimizer_D_A, optimizer_D_B=optimizer_D_B,
        adversarial_loss=adversarial_loss, cycle_loss=cycle_loss, pixelwise_loss=pixelwise_loss,
        device=device, model_fold_dir=model_fold_dir, cfg_trainer=cfg['trainer'],
    )

    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)


    # Test finale
    mse = nn.MSELoss().to(device) # mse --> adversarial loss
    psnr = PeakSignalNoiseRatio().to(device)
    mse_t, psnr_t, vif_t, ssim_t = util_model.test_4metrics(G_AB=G_AB, loader_test=data_loaders['test'],
                                                            mse_metric=mse, psnr_metric=psnr,
                                                            device=device, outputs_dir=outputs_fold_dir,
                                                            save_output=True)
    print(f"FOLD {fold} → MSE: {mse_t:.4f}, PSNR: {psnr_t:.4f}, VIF: {vif_t:.4f}, SSIM: {ssim_t:.4f}")

    # Update report
    results_mse["%s MSE" % str(fold)].append(mse_t)
    results_psnr["%s PSNR" % str(fold)].append(psnr_t)
    results_vif["%s VIF" % str(fold)].append(vif_t)
    results_ssim["%s SSIM" % str(fold)].append(ssim_t)

    # Save Results MSE
    results_mse_frame = pd.DataFrame(results_mse)
    results_mse_frame.insert(loc=0, column='std MSE', value=results_mse_frame[metric_cols_mse].std(axis=1))
    results_mse_frame.insert(loc=0, column='mean MSE', value=results_mse_frame[metric_cols_mse].mean(axis=1))
    results_mse_frame.insert(loc=0, column='model', value=model_name)
    results_mse_frame.to_excel(report_file_mse, index=False)

    # Save Results PSNR
    results_psnr_frame = pd.DataFrame(results_psnr)
    results_psnr_frame.insert(loc=0, column='std PSNR', value=results_psnr_frame[metric_cols_psnr].std(axis=1))
    results_psnr_frame.insert(loc=0, column='mean PSNR', value=results_psnr_frame[metric_cols_psnr].mean(axis=1))
    results_psnr_frame.insert(loc=0, column='model', value=model_name)
    results_psnr_frame.to_excel(report_file_psnr, index=False)

    # Save Results VIF
    results_vif_frame = pd.DataFrame(results_vif)
    results_vif_frame.insert(loc=0, column='std VIF', value=results_vif_frame[metric_cols_vif].std(axis=1))
    results_vif_frame.insert(loc=0, column='mean VIF', value=results_vif_frame[metric_cols_vif].mean(axis=1))
    results_vif_frame.insert(loc=0, column='model', value=model_name)
    results_vif_frame.to_excel(report_file_vif, index=False)

    # Save Results SSIM
    results_ssim_frame = pd.DataFrame(results_ssim)
    results_ssim_frame.insert(loc=0, column='std SSIM', value=results_ssim_frame[metric_cols_ssim].std(axis=1))
    results_ssim_frame.insert(loc=0, column='mean SSIM', value=results_ssim_frame[metric_cols_ssim].mean(axis=1))
    results_ssim_frame.insert(loc=0, column='model', value=model_name)
    results_ssim_frame.to_excel(report_file_ssim, index=False)

