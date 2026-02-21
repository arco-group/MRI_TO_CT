import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])
import torch
import torch.nn as nn
import torch.optim as optim
from src.model.Pix2Pix.generator_model import Generator
from src.model.Pix2Pix.discriminator_model import Discriminator
import src.model.Pix2Pix.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data
from torchmetrics import PeakSignalNoiseRatio
import os
import yaml
import ssl
import collections
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from sys import platform

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# Configuration file
if platform == 'win32': #windows
    args = {}
    args['cfg_file'] = "./configs/pix2pix_train.yaml"

    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)
else:
    args ={} #others
    args['cfg_file'] = ".../configs/pix2pix_train.yaml" #select the path where the file will be saved
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)

# Parameters
exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
learning_rate = cfg['trainer']['optimizer']['lr']

fold_env = os.environ.get("FOLD_IDX", None)
if fold_env is not None:
    fold_list = [int(fold_env)]
    print(f"*** Execution only for the fold:  {fold_list[0]} ")
else:
    cv = cfg['data']['cv']
    fold_list = list(range(cv))
    print(f"*** Execution for all folds: {fold_list} ***")

# Device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(device)


# Files and Directories
fold_dir = cfg['data']['fold_dir']
model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)
report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)
outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

for fold in fold_list:
    print("Fold %d" % fold)
    # Dir
    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_DISC'])):
        print("Skipping Fold %d" % fold)
        continue
    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(outputs_fold_dir)

    # Data Loaders 2D
    fold_data = {step: pd.read_csv(os.path.join(fold_dir, str(fold), '%s.csv' % step), index_col='id_slice') for step in ['train', 'val', 'test']}
    datasets = {step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}

    data_loaders = {
        'train': torch.utils.data.DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'val': torch.utils.data.DataLoader(datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False,num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'test': torch.utils.data.DataLoader(datasets['test'], batch_size=1, shuffle=False, num_workers=num_workers,worker_init_fn=util_data.seed_worker)}

    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name'
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size'
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs'
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim'
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)


    # Initialize the models

    disc = Discriminator(in_channels=1).to(device).float()
    gen = Generator(in_channels=1, features=64).to(device).float()

    opt_disc = optim.Adam(disc.parameters(), lr=learning_rate, betas=(0.5, 0.999), )
    opt_gen = optim.Adam(gen.parameters(), lr=learning_rate, betas=(0.5, 0.999))

    BCE = nn.BCEWithLogitsLoss()
    L1_LOSS = nn.L1Loss()

    early_stopping_criterion = nn.L1Loss().to(device)

    # Train model
    gen, disc, history = util_model.train_pix2pix(gen=gen, disc=disc, data_loaders=data_loaders, early_stopping_criterion=early_stopping_criterion, opt_disc=opt_disc, opt_gen=opt_gen, L1=L1_LOSS, bce=BCE, model_fold_dir=model_fold_dir, cfg_trainer=cfg['trainer'], device=device)

    # Plot Training
    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    mse = nn.MSELoss()
    psnr = PeakSignalNoiseRatio().to(device)

    mse_test_results, psnr_test_results, vif_test_results, ssim_test_results = util_model.test_4metrics(gen=gen, loader=data_loaders['test'], mse_metric=mse, psnr_metric=psnr, device=device, outputs_dir=outputs_fold_dir, save_output=True)
    print(f"MSE: {mse_test_results:.4f}, PSNR: {psnr_test_results:.4f}, VIF: {vif_test_results:.4f}, SSIM: {ssim_test_results:.4f}")





