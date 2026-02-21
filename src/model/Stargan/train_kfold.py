import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])
import os
import ssl
import yaml
import torch
import torch.nn as nn
import torch.optim as optim
import pandas as pd
import collections
from torch.utils.data import DataLoader
from torchmetrics import PeakSignalNoiseRatio
from openpyxl import Workbook, load_workbook
from sys import platform
from src.model.Stargan.generator_model import GeneratorResNet
from src.model.Stargan.discriminator_model import Discriminator
import src.model.Stargan.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# Config
if platform == 'win32':
    cfg_path = "./configs/stargan_train.yaml"
else:
    cfg_path = ".../configs/stargan_train.yaml"

with open(cfg_path) as f:
    cfg = yaml.load(f, Loader=yaml.FullLoader)

# Parameters
exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']

fold_env = os.environ.get("FOLD_IDX", None)
if fold_env is not None:
    fold_list = [int(fold_env)]
    print(f"*** Execution only for the fold {fold_list[0]}  ***")
else:
    cv = cfg['data']['cv']
    fold_list = list(range(cv))
    print(f"*** execution for all folds: {fold_list} ***")

# Device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(f"Device: {device}")

# Files & directories
fold_dir = os.path.join(cfg['data']['fold_dir'])

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)





for fold in fold_list:
    print(f"\n=== Fold {fold} ===")

    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_G'])):
        print(f"Skipping Fold {fold} (checkpoint)")
        continue

    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    util_general.create_dir(outputs_fold_dir)



    fold_data = {
        step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col='id_slice')
        for step in ['train', 'val', 'test']
    }

    datasets = {
        step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True)
        for step in ['train', 'val', 'test']
    }

    data_loaders = {
        'train': DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'],
                            shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'val':   DataLoader(datasets['val'], batch_size=cfg['data']['batch_size'],
                            shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'test':  DataLoader(datasets['test'], batch_size=1,
                            shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
    }

    print(
        f"Fold {fold} - #train: {len(datasets['train'])}, "
        f"#val: {len(datasets['val'])}, #test: {len(datasets['test'])}"
    )

    # info.xlsx
    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name'
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size'
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs'
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim'
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)

    # Models
    img_dim = cfg['data']['img_dim']
    img_shape = (1, img_dim, img_dim)
    c_dim = 2  # MRI/CT

    generator = GeneratorResNet(
        img_shape=img_shape,
        res_blocks=cfg['model']['residual_blocks'],
        c_dim=c_dim
    ).to(device)

    discriminator = Discriminator(
        img_shape=img_shape,
        c_dim=c_dim,
        n_strided=cfg['model']['n_strided']
    ).to(device)


    lr = cfg['trainer']['optimizer']['lr']
    opt_G = optim.Adam(generator.parameters(), lr=lr, betas=(0.5, 0.999))
    opt_D = optim.Adam(discriminator.parameters(), lr=lr, betas=(0.5, 0.999))

    # Loss & lambdas
    l1_cycle = nn.L1Loss().to(device)
    lambdas = {
        "lambda_cls": cfg['trainer']['lambda_cls'],
        "lambda_rec": cfg['trainer']['lambda_rec'],
        "lambda_gp":  cfg['trainer']['lambda_gp'],
    }
    n_critic = cfg['trainer']['n_critic']

    # Train
    generator, discriminator, history = util_model.train_stargan(
        generator=generator,
        discriminator=discriminator,
        data_loaders=data_loaders,
        opt_G=opt_G,
        opt_D=opt_D,
        l1_cycle=l1_cycle,
        lambdas=lambdas,
        n_critic=n_critic,
        model_fold_dir=model_fold_dir,
        cfg_trainer=cfg['trainer'],
        device=device,
    )

    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # Test (MRI->CT)
    mse_t, psnr_t, vif_t, ssim_t = util_model.test_4metrics_stargan(
        generator=generator,
        loader_test=data_loaders['test'],
        device=device,
        outputs_dir=outputs_fold_dir,
        save_output=True,
    )

    print(f"Fold {fold} - MSE: {mse_t:.4f}, PSNR: {psnr_t:.4f}, VIF: {vif_t:.4f}, SSIM: {ssim_t:.4f}")



