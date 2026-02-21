import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])
import torch
from torch.utils.data import DataLoader
import torch.nn as nn
import torch.optim as optim
import os
import yaml
import ssl
import pandas as pd
import collections
from torchmetrics import PeakSignalNoiseRatio
from src.model.cyclegan.discriminator_model import Discriminator
from src.model.cyclegan.generator_model import Generator
import src.utils.util_general as util_general
import src.utils.util_data as util_data
import src.model.cyclegan.util_model as util_model
from openpyxl import Workbook
from openpyxl import load_workbook
from sys import platform

torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# Configuration file

if platform == 'win32': # su windows
    args = {}
    args['cfg_file'] = "./configs/cyclegan_train.yaml"
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)
else: # su alvis
    args = {}
    args['cfg_file'] = ".../configs/cyclegan_train.yaml"
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)


# Parameters
exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
learning_rate = cfg['trainer']['optimizer']['lr']

fold_env = os.environ.get("FOLD_IDX", None)
if fold_env is not None:
    fold_list = [int(fold_env)]
    print(f"*** Esecuzione solo per la fold {fold_list[0]} (da variabile d'ambiente FOLD_IDX) ***")
else:
    cv = cfg['data']['cv']
    fold_list = list(range(cv))
    print(f"*** Esecuzione per tutte le fold: {fold_list} ***")


# Device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(device)


fold_dir = os.path.join(cfg['data']['fold_dir'])

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

# CV
results_mse = collections.defaultdict(lambda: [])
metric_cols_mse = []
results_psnr = collections.defaultdict(lambda: [])
metric_cols_psnr = []
results_vif = collections.defaultdict(lambda: [])
metric_cols_vif = []
results_ssim = collections.defaultdict(lambda: [])
metric_cols_ssim = []

for fold in fold_list:
    print('Fold %s' % str(fold))

    # Dir
    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_CRITIC_LE'])):
        print("Skipping Fold %d" % fold)
        continue
    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(outputs_fold_dir)


    # Data Loaders
    fold_data = {step: pd.read_csv(os.path.join(fold_dir, str(fold), '%s.csv' % step), index_col='id_slice') for step in ['train', 'val', 'test']}
    datasets = {step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}
    print(
        f"Fold {fold} - #train: {len(datasets['train'])}, "
        f"#val: {len(datasets['val'])}, #test: {len(datasets['test'])}"
    )
    data_loaders = {
        'train': torch.utils.data.DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'val': torch.utils.data.DataLoader(datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'test': torch.utils.data.DataLoader(datasets['test'], batch_size=1, shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker)}


    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)

        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name'
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size'
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs'
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim'
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)


    disc_y = Discriminator(in_channels=1).to(device)
    disc_x = Discriminator(in_channels=1).to(device)
    gen_x = Generator(img_channels=1, num_residuals=9).to(device)
    gen_y = Generator(img_channels=1, num_residuals=9).to(device)


    opt_disc = optim.Adam(
        list(disc_y.parameters()) + list(disc_x.parameters()),
        lr=learning_rate,
        betas=(0.5, 0.999),
    )

    opt_gen = optim.Adam(
        list(gen_x.parameters()) + list(gen_y.parameters()),
        lr=learning_rate,
        betas=(0.5, 0.999),
    )

    L1 = nn.L1Loss().to(device)
    mse = nn.MSELoss().to(device)
    early_stopping_criterion = nn.L1Loss().to(device)

    # Train model
    gen_y, gen_x, disc_y, disc_x, history = util_model.train_cycle_gan(gen_REC=gen_y, gen_LE=gen_x, disc_REC=disc_y, disc_LE=disc_x,
                                                                             data_loaders=data_loaders, early_stopping_criterion=early_stopping_criterion,
                                                                             opt_disc=opt_disc, opt_gen=opt_gen, L1=L1, mse=mse,
                                                                             model_fold_dir=model_fold_dir, cfg_trainer=cfg['trainer'], device=device)


    # Plot Training
    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # Test model
    mse = nn.MSELoss()
    psnr = PeakSignalNoiseRatio().to(device)

    mse_test_results, psnr_test_results, vif_test_results, ssim_test_results = util_model.test_4metrics(gen_y
                                                                                                        , loader_train=False, loader_val=False, loader_test=data_loaders['test'], mse_metric=mse, psnr_metric=psnr, device=device, outputs_dir=outputs_fold_dir, save_output=True)
    print('MSE:', mse_test_results, 'PSNR:', psnr_test_results, 'VIF:', vif_test_results, 'SSIM:', ssim_test_results)


