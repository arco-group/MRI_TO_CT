import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])
import torch
from torch import nn
import torch.optim as optim
from torch.utils.data import DataLoader
import os
import yaml
import ssl
import pandas as pd
import collections
from torchmetrics import PeakSignalNoiseRatio
from sys import platform as sys_platform
from src.model.Bicyclegan.generator_model import Generator, Encoder, weights_init_normal
from src.model.Bicyclegan.discriminator_model import MultiDiscriminator
import src.model.Bicyclegan.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data
from openpyxl import Workbook
from openpyxl import load_workbook

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

#Reading and configuration
if sys_platform == 'win32':
    args = {}
    args['cfg_file'] = "./configs/bicyclegan.yaml"
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)
else:
    args = {}
    args['cfg_file'] = ".../configs/bicyclegan_train.yaml"
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)

# Parameters
exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
learning_rate = cfg['trainer']['optimizer']['lr']

fold_env = os.environ.get("FOLD_IDX", None)
if fold_env is not None:
    fold_list = [int(fold_env)]
    print(f"*** Fold execution {fold_list[0]} ***")
else:
    cv = cfg['data']['cv']
    fold_list = list(range(cv))
    print(f"*** Execution of all folds: {fold_list} ***")


# Device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(f"Device:  {device}")

# Files & directories
fold_dir = cfg['data']['fold_dir']

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

# CV results
results_mse = collections.defaultdict(lambda: [])
metric_cols_mse = []
results_psnr = collections.defaultdict(lambda: [])
metric_cols_psnr = []
results_vif = collections.defaultdict(lambda: [])
metric_cols_vif = []
results_ssim = collections.defaultdict(lambda: [])
metric_cols_ssim = []

for fold in fold_list:
    print(f"\n===== Fold {fold} =====")

    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)


    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_GEN'])):
        print(f"Skipping Fold {fold} (checkpoint found)")
        continue

    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(outputs_fold_dir)


    # Data Loaders: train/val/test da CSV (5-fold)
    fold_data = {
        step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col='id_slice') for step in ['train', 'val', 'test']
    }
    datasets = { step: util_data.ImgDataset( data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}
    data_loaders = {
        'train': torch.utils.data.DataLoader(
            datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'val': torch.utils.data.DataLoader( datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'test': torch.utils.data.DataLoader(datasets['test'], batch_size=1, shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
    }


    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name';
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size';
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs';
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim';
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)


    img_dim = cfg['data']['img_dim']
    latent_dim = cfg['model']['latent_dim']

    generator = Generator( latent_dim=latent_dim, img_height=img_dim, img_width=img_dim, in_channels=1, out_channels=1, out_activation='sigmoid').to(device)
    encoder = Encoder(latent_dim=latent_dim, in_channels=1).to(device)
    D_VAE = MultiDiscriminator(in_channels=1).to(device)
    D_LR = MultiDiscriminator(in_channels=1).to(device)


    # Ottimizzatori
    opt_E = optim.Adam(encoder.parameters(), lr=learning_rate, betas=(0.5, 0.999))
    opt_G = optim.Adam(generator.parameters(), lr=learning_rate, betas=(0.5, 0.999))
    opt_D_VAE = optim.Adam(D_VAE.parameters(), lr=learning_rate, betas=(0.5, 0.999))
    opt_D_LR = optim.Adam(D_LR.parameters(), lr=learning_rate, betas=(0.5, 0.999))

    # Loss & lambdas
    l1_pix = nn.L1Loss().to(device)
    early_stopping_criterion = nn.L1Loss().to(device)

    lambdas = {
        "lambda_pixel": cfg['trainer']['LAMBDA_PIXEL'],
        "lambda_latent": cfg['trainer']['LAMBDA_LATENT'],
        "lambda_kl": cfg['trainer']['LAMBDA_KL'],
    }

    # Train
    generator, encoder, D_VAE, D_LR, history = util_model.train_bicyclegan(
        generator=generator,
        encoder=encoder,
        D_VAE=D_VAE,
        D_LR=D_LR,
        data_loaders=data_loaders,
        opt_E=opt_E,
        opt_G=opt_G,
        opt_D_VAE=opt_D_VAE,
        opt_D_LR=opt_D_LR,
        l1_pix=l1_pix,
        lambdas=lambdas,
        early_stopping_criterion=early_stopping_criterion,
        model_fold_dir=model_fold_dir,
        cfg_trainer=cfg['trainer'],
        device=device,
    )

    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # Test
    mse_metric = nn.MSELoss()
    psnr_metric = PeakSignalNoiseRatio().to(device)

    mse_test, psnr_test, vif_test, ssim_test = util_model.test_4metrics(
        generator=generator,
        encoder=encoder,
        loader=data_loaders['test'],
        mse_metric=mse_metric,
        psnr_metric=psnr_metric,
        device=device,
        outputs_dir=outputs_fold_dir,
        save_output=True,
    )
    print(f"Fold {fold} - MSE: {mse_test:.4f}, PSNR: {psnr_test:.4f}, VIF: {vif_test:.4f}, SSIM: {ssim_test:.4f}")


