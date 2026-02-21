import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])
import torch
from torch import nn
import torch.optim as optim
from torch.utils.data import DataLoader
import os
import yaml
import ssl
import pandas as pd
import collections
from torchmetrics import PeakSignalNoiseRatio
from sys import platform as sys_platform
from src.model.Bicyclegan.generator_model import Generator, Encoder, weights_init_normal
from src.model.Bicyclegan.discriminator_model import MultiDiscriminator
import src.model.Bicyclegan.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data
from openpyxl import Workbook
from openpyxl import load_workbook

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

#Lettura e configurazione
if sys_platform == 'win32':
    args = {}
    args['cfg_file'] = "./configs/bicyclegan.yaml"
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)
else:
    args = {}
    args['cfg_file'] = "/mimer/NOBACKUP/groups/snic2022-5-277/rofena/Progetto_Alessandro/configs/bicyclegan_train.yaml"
    with open(args['cfg_file']) as file:
        cfg = yaml.load(file, Loader=yaml.FullLoader)

# Parameters
exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
learning_rate = cfg['trainer']['optimizer']['lr']
cv = cfg['data']['cv']
fold_list = list(range(cv))

# Device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(f"Device:  {device}")

# Files & directories
fold_dir = cfg['data']['fold_dir'] #da dove prende il dataset

model_dir = os.path.join(cfg['data']['model_dir'], exp_name) #dove carica i pesi dei modelli
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name) #dove carica i risultati di output e loss
util_general.create_dir(report_dir)

report_file_mse = os.path.join(report_dir, 'report_mse.xlsx')
report_file_psnr = os.path.join(report_dir, 'report_psnr.xlsx')
report_file_vif = os.path.join(report_dir, 'report_vif.xlsx')
report_file_ssim = os.path.join(report_dir, 'report_ssim.xlsx')

plot_training_dir = os.path.join(report_dir, "training") #dove saranno le loss
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs") #qui salverà le immagini di output
util_general.create_dir(outputs_dir)

# CV results
results_mse = collections.defaultdict(lambda: [])
metric_cols_mse = []
results_psnr = collections.defaultdict(lambda: [])
metric_cols_psnr = []
results_vif = collections.defaultdict(lambda: [])
metric_cols_vif = []
results_ssim = collections.defaultdict(lambda: [])
metric_cols_ssim = []

for fold in fold_list:
    print(f"\n===== Fold {fold} =====")

    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)

    # se esiste già il checkpoint del generator, puoi saltare il fold (opzionale)
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_GEN'])):
        print(f"Skipping Fold {fold} (checkpoint found)")
        continue

    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(outputs_fold_dir)

    # colonne report
    metric_cols_mse.append(f"{fold} MSE")
    metric_cols_psnr.append(f"{fold} PSNR")
    metric_cols_vif.append(f"{fold} VIF")
    metric_cols_ssim.append(f"{fold} SSIM")

    # Data Loaders: train/val/test da CSV (5-fold)
    fold_data = {
        step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col='id_slice') for step in ['train', 'val', 'test']
    }#legge il dataset corrispondente e lo salva in fold_data
    datasets = { step: util_data.ImgDataset( data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}
    data_loaders = {#caricamento dei dati
        'train': torch.utils.data.DataLoader(
            datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'val': torch.utils.data.DataLoader( datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'test': torch.utils.data.DataLoader(datasets['test'], batch_size=1, shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
    }

    # Crea info.xlsx solo al primo fold
    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name';
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size';
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs';
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim';
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)

    # Inizializzo i Modelli
    img_dim = cfg['data']['img_dim']
    latent_dim = cfg['model']['latent_dim']

    generator = Generator( latent_dim=latent_dim, img_height=img_dim, img_width=img_dim, in_channels=1, out_channels=1, out_activation='sigmoid').to(device)
    encoder = Encoder(latent_dim=latent_dim, in_channels=1).to(device)
    D_VAE = MultiDiscriminator(in_channels=1).to(device)
    D_LR = MultiDiscriminator(in_channels=1).to(device)

    # init pesi
   # generator.apply(weights_init_normal)
    #encoder.apply(weights_init_normal)
    #D_VAE.apply(weights_init_normal)
    #D_LR.apply(weights_init_normal)

    # Ottimizzatori
    opt_E = optim.Adam(encoder.parameters(), lr=learning_rate, betas=(0.5, 0.999))
    opt_G = optim.Adam(generator.parameters(), lr=learning_rate, betas=(0.5, 0.999))
    opt_D_VAE = optim.Adam(D_VAE.parameters(), lr=learning_rate, betas=(0.5, 0.999))
    opt_D_LR = optim.Adam(D_LR.parameters(), lr=learning_rate, betas=(0.5, 0.999))

    # Loss & lambdas
    l1_pix = nn.L1Loss().to(device)
    early_stopping_criterion = nn.L1Loss().to(device)

    lambdas = {
        "lambda_pixel": cfg['trainer']['LAMBDA_PIXEL'],
        "lambda_latent": cfg['trainer']['LAMBDA_LATENT'],
        "lambda_kl": cfg['trainer']['LAMBDA_KL'],
    }

    # Train
    generator, encoder, D_VAE, D_LR, history = util_model.train_bicyclegan(
        generator=generator,
        encoder=encoder,
        D_VAE=D_VAE,
        D_LR=D_LR,
        data_loaders=data_loaders,
        opt_E=opt_E,
        opt_G=opt_G,
        opt_D_VAE=opt_D_VAE,
        opt_D_LR=opt_D_LR,
        l1_pix=l1_pix,
        lambdas=lambdas,
        early_stopping_criterion=early_stopping_criterion,
        model_fold_dir=model_fold_dir,
        cfg_trainer=cfg['trainer'],
        device=device,
    )

    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # Test
    mse_metric = nn.MSELoss()
    psnr_metric = PeakSignalNoiseRatio().to(device)

    mse_test, psnr_test, vif_test, ssim_test = util_model.test_4metrics(
        generator=generator,
        encoder=encoder,
        loader=data_loaders['test'],
        mse_metric=mse_metric,
        psnr_metric=psnr_metric,
        device=device,
        outputs_dir=outputs_fold_dir,
        save_output=True,
    )
    print(f"Fold {fold} - MSE: {mse_test:.4f}, PSNR: {psnr_test:.4f}, VIF: {vif_test:.4f}, SSIM: {ssim_test:.4f}")

    # Update report
    results_mse[f"{fold} MSE"].append(mse_test)
    results_psnr[f"{fold} PSNR"].append(psnr_test)
    results_vif[f"{fold} VIF"].append(vif_test)
    results_ssim[f"{fold} SSIM"].append(ssim_test)

    # Save Results MSE
    results_mse_frame = pd.DataFrame(results_mse)
    results_mse_frame.insert(loc=0, column='std MSE', value=results_mse_frame[metric_cols_mse].std(axis=1))
    results_mse_frame.insert(loc=0, column='mean MSE', value=results_mse_frame[metric_cols_mse].mean(axis=1))
    results_mse_frame.insert(loc=0, column='model', value=model_name)
    results_mse_frame.to_excel(report_file_mse, index=False)

    # Save Results PSNR
    results_psnr_frame = pd.DataFrame(results_psnr)
    results_psnr_frame.insert(loc=0, column='std PSNR', value=results_psnr_frame[metric_cols_psnr].std(axis=1))
    results_psnr_frame.insert(loc=0, column='mean PSNR', value=results_psnr_frame[metric_cols_psnr].mean(axis=1))
    results_psnr_frame.insert(loc=0, column='model', value=model_name)
    results_psnr_frame.to_excel(report_file_psnr, index=False)

    # Save Results VIF
    results_vif_frame = pd.DataFrame(results_vif)
    results_vif_frame.insert(loc=0, column='std VIF', value=results_vif_frame[metric_cols_vif].std(axis=1))
    results_vif_frame.insert(loc=0, column='mean VIF', value=results_vif_frame[metric_cols_vif].mean(axis=1))
    results_vif_frame.insert(loc=0, column='model', value=model_name)
    results_vif_frame.to_excel(report_file_vif, index=False)

    # Save Results SSIM
    results_ssim_frame = pd.DataFrame(results_ssim)
    results_ssim_frame.insert(loc=0, column='std SSIM', value=results_ssim_frame[metric_cols_ssim].std(axis=1))
    results_ssim_frame.insert(loc=0, column='mean SSIM', value=results_ssim_frame[metric_cols_ssim].mean(axis=1))
    results_ssim_frame.insert(loc=0, column='model', value=model_name)
    results_ssim_frame.to_excel(report_file_ssim, index=False)
