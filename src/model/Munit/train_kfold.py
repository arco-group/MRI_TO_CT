import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
import os
import yaml
import ssl
import pandas as pd
import collections
from sys import platform
from openpyxl import Workbook
from openpyxl import load_workbook
from torchmetrics import PeakSignalNoiseRatio
from src.model.Munit.generator_model import Encoder, Decoder
from src.model.Munit.discriminator_model import MultiDiscriminator
import src.model.Munit.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# Config
if platform == 'win32':
    args = {}
    args['cfg_file'] = "./configs/munit_train.yaml"
    with open(args['cfg_file']) as f:
        cfg = yaml.load(f, Loader=yaml.FullLoader)
else:
    args = {}
    args['cfg_file'] = "./configs/munit_train.yaml"
    with open(args['cfg_file']) as f:
        cfg = yaml.load(f, Loader=yaml.FullLoader)

#Parameters
exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
lr = cfg['trainer']['optimizer']['lr']
# Parameters
exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
learning_rate = cfg['trainer']['optimizer']['lr']

fold_env = os.environ.get("FOLD_IDX", None)
if fold_env is not None:
    fold_list = [int(fold_env)]
    print(f"*** Esecution only for fold {fold_list[0]} (FOLD_IDX) ***")
else:
    cv = cfg['data']['cv']
    fold_list = list(range(cv))
    print(f"*** Esecution on all folds: {fold_list} ***")


# Device
device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print(f"Device:  {device}")

# Files & directories
fold_dir = os.path.join(cfg['data']['fold_dir'])
model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

for fold in fold_list:
    print(f"\n=== Fold {fold} ===")

    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_ENC1'])):
        print(f"Skipping Fold {fold} (checkpoint found)")
        continue


    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(outputs_fold_dir)



    fold_data = {
        step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col='id_slice') for step in ['train', 'val', 'test']
    }

    datasets = {
        step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True) for step in ['train', 'val', 'test']}

    data_loaders = {
        'train': DataLoader(datasets['train'], batch_size=cfg['data']['batch_size'], shuffle=True, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'val':   DataLoader(datasets['val'], batch_size=cfg['data']['batch_size'], shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker),
        'test':  DataLoader(datasets['test'], batch_size=1, shuffle=False, num_workers=num_workers, worker_init_fn=util_data.seed_worker)
    }

    print(
        f"Fold {fold} - #train: {len(datasets['train'])}, "
        f"#val: {len(datasets['val'])}, #test: {len(datasets['test'])}"
    )


    if fold == 0:
        wb = Workbook()
        filepath = os.path.join(model_dir, "info.xlsx")
        wb.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active
        sheet['A1'] = 'exp_name';
        sheet['A2'] = exp_name
        sheet['B1'] = 'batch size';
        sheet['B2'] = cfg['data']['batch_size']
        sheet['C1'] = 'epochs';
        sheet['C2'] = cfg['trainer']['max_epochs']
        sheet['D1'] = 'img_dim';
        sheet['D2'] = cfg['data']['img_dim']
        wb.save(filepath)


    dim = cfg['model']['dim']
    n_down = cfg['model']['n_downsample']
    n_res = cfg['model']['n_residual']
    style_dim = cfg['model']['style_dim']

    Enc1 = Encoder(in_channels=1, dim=dim, n_residual=n_res, n_downsample=n_down, style_dim=style_dim).to(device)
    Dec1 = Decoder(out_channels=1, dim=dim, n_residual=n_res, n_upsample=n_down, style_dim=style_dim).to(device)
    Enc2 = Encoder(in_channels=1, dim=dim, n_residual=n_res, n_downsample=n_down, style_dim=style_dim).to(device)
    Dec2 = Decoder(out_channels=1, dim=dim, n_residual=n_res, n_upsample=n_down, style_dim=style_dim).to(device)
    D1   = MultiDiscriminator(in_channels=1).to(device)
    D2   = MultiDiscriminator(in_channels=1).to(device)


    opt_G  = optim.Adam(
        list(Enc1.parameters()) + list(Dec1.parameters()) +
        list(Enc2.parameters()) + list(Dec2.parameters()),
        lr=lr, betas=(0.5, 0.999)
    )
    opt_D1 = optim.Adam(D1.parameters(), lr=lr, betas=(0.5, 0.999))
    opt_D2 = optim.Adam(D2.parameters(), lr=lr, betas=(0.5, 0.999))

    # Loss & lambdas
    l1_recon = nn.L1Loss().to(device)

    lambdas = {
        "LAMBDA_GAN":     cfg['trainer']['LAMBDA_GAN'],
        "LAMBDA_ID":      cfg['trainer']['LAMBDA_ID'],
        "LAMBDA_STYLE":   cfg['trainer']['LAMBDA_STYLE'],
        "LAMBDA_CONTENT": cfg['trainer']['LAMBDA_CONTENT'],
        "LAMBDA_CYCLE":   cfg['trainer']['LAMBDA_CYCLE'],
        "STYLE_DIM":      style_dim,
    }

    # Train
    Enc1, Dec1, Enc2, Dec2, D1, D2, history = util_model.train_munit(
        Enc1, Dec1, Enc2, Dec2, D1, D2,
        data_loaders=data_loaders,
        opt_G=opt_G, opt_D1=opt_D1, opt_D2=opt_D2,
        l1_recon=l1_recon,
        lambdas=lambdas,
        model_fold_dir=model_fold_dir,
        cfg_trainer=cfg['trainer'],
        device=device,
    )

    util_model.plot_training(history=history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # Test (MRI->CT)
    mse, psnr, vif, ssim = util_model.test_4metrics(
        Enc1=Enc1, Dec2=Dec2,
        loader_test=data_loaders['test'],
        device=device,
        outputs_dir=outputs_fold_dir,
        save_output=True,
    )

    print(f"Fold {fold} - MSE: {mse:.4f}, PSNR: {psnr:.4f}, VIF: {vif:.4f}, SSIM: {ssim:.4f}")


