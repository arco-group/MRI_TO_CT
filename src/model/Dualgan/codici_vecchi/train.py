import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(["./"])

import os
import yaml
import ssl
import torch
import torch.nn as nn
import torch.optim as optim
import pandas as pd
import collections
from torchmetrics import PeakSignalNoiseRatio
from sys import platform

from src.model.Dualgan.discriminator_model import Discriminator
from src.model.Dualgan.generator_model import Generator, weights_init_normal
import src.model.Dualgan.util_model as util_model
import src.utils.util_general as util_general
import src.utils.util_data as util_data
from openpyxl import Workbook, load_workbook

torch.backends.cudnn.benchmark = True
torch.cuda.empty_cache()
os.environ['TORCH_HOME'] = './model/cnn/pretrained'
ssl._create_default_https_context = ssl._create_unverified_context

# Config
if platform == 'win32':
    cfg_path = "./configs/dualgan_train.yaml"
else:
    cfg_path = "/mimer/NOBACKUP/groups/snic2022-5-277/rofena/Progetto_Alessandro/configs/dualgan_train.yaml"

with open(cfg_path) as f:
    cfg = yaml.load(f, Loader=yaml.FullLoader)

exp_name = cfg['exp_name']
print(exp_name)
model_name = cfg['model']['model_name']
cv = cfg['data']['cv']
fold_list = list(range(cv))
# fold selection via env

device = torch.device(cfg['device']['cuda_device'] if torch.cuda.is_available() else "cpu")
num_workers = 0 if device.type == "cpu" else cfg['device']['gpu_num_workers']
print("Device:", device)

fold_dir = os.path.join(cfg['data']['fold_dir'])

model_dir = os.path.join(cfg['data']['model_dir'], exp_name)
util_general.create_dir(model_dir)

report_dir = os.path.join(cfg['data']['report_dir'], exp_name)
util_general.create_dir(report_dir)

report_file_mse  = os.path.join(report_dir, 'report_mse.xlsx')
report_file_psnr = os.path.join(report_dir, 'report_psnr.xlsx')
report_file_vif  = os.path.join(report_dir, 'report_vif.xlsx')
report_file_ssim = os.path.join(report_dir, 'report_ssim.xlsx')

plot_training_dir = os.path.join(report_dir, "training")
util_general.create_dir(plot_training_dir)

outputs_dir = os.path.join(report_dir, "outputs")
util_general.create_dir(outputs_dir)

results_mse  = collections.defaultdict(lambda: [])
results_psnr = collections.defaultdict(lambda: [])
results_vif  = collections.defaultdict(lambda: [])
results_ssim = collections.defaultdict(lambda: [])

metric_cols_mse  = []
metric_cols_psnr = []
metric_cols_vif  = []
metric_cols_ssim = []

for fold in fold_list:
    print(f"\n========== Fold {fold} ==========")

    model_fold_dir = os.path.join(model_dir, str(fold))
    util_general.create_dir(model_fold_dir)

    # skip if already trained
    if os.path.exists(os.path.join(model_fold_dir, cfg['trainer']['CHECKPOINT_G_AB'])):
        print(f"Skipping Fold {fold} (checkpoint found)")
        continue

    plot_training_fold_dir = os.path.join(plot_training_dir, str(fold))
    outputs_fold_dir = os.path.join(outputs_dir, str(fold))
    util_general.create_dir(plot_training_fold_dir)
    util_general.create_dir(outputs_fold_dir)

    metric_cols_mse.append(f"{fold} MSE")
    metric_cols_psnr.append(f"{fold} PSNR")
    metric_cols_vif.append(f"{fold} VIF")
    metric_cols_ssim.append(f"{fold} SSIM")

    # --------------------
    # Data loaders ALE (CSV kfold)
    # --------------------
    fold_data = {
        step: pd.read_csv(os.path.join(fold_dir, str(fold), f"{step}.csv"), index_col="id_slice")
        for step in ["train", "val", "test"]
    }
    datasets = {
        step: util_data.ImgDataset(data=fold_data[step], cfg_data=cfg['data'], step=step, do_augmentation=True)
        for step in ["train", "val", "test"]
    }

    pf = cfg['device'].get("prefetch_factor", 2) if num_workers > 0 else None
    data_loaders = {
        "train": torch.utils.data.DataLoader(
            datasets["train"],
            batch_size=cfg['data']['batch_size'],
            shuffle=True,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False,
            prefetch_factor=pf if num_workers > 0 else None,
        ),
        "val": torch.utils.data.DataLoader(
            datasets["val"],
            batch_size=cfg['data']['batch_size'],
            shuffle=False,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False,
            prefetch_factor=pf if num_workers > 0 else None,
        ),
        "test": torch.utils.data.DataLoader(
            datasets["test"],
            batch_size=1,
            shuffle=False,
            num_workers=num_workers,
            worker_init_fn=util_data.seed_worker,
            pin_memory=True,
            persistent_workers=True if num_workers > 0 else False,
            prefetch_factor=pf if num_workers > 0 else None,
        ),
    }

    # info.xlsx solo al primo fold
    if fold == 0:
        filepath = os.path.join(model_dir, "info.xlsx")
        wb = Workbook(); wb.save(filepath)
        wb = load_workbook(filepath)
        s = wb.active
        s['A1'] = 'exp_name'; s['A2'] = exp_name
        s['B1'] = 'batch size'; s['B2'] = cfg['data']['batch_size']
        s['C1'] = 'epochs'; s['C2'] = cfg['trainer']['max_epochs']
        s['D1'] = 'img_dim'; s['D2'] = cfg['data']['img_dim']
        wb.save(filepath)

    # --------------------
    # Models
    # --------------------
    G_AB = Generator(channels=1).to(device)
    G_BA = Generator(channels=1).to(device)
    D_A  = Discriminator(in_channels=1).to(device)
    D_B  = Discriminator(in_channels=1).to(device)

    G_AB.apply(weights_init_normal)
    G_BA.apply(weights_init_normal)
    D_A.apply(weights_init_normal)
    D_B.apply(weights_init_normal)

    # Optimizers
    lr = cfg['trainer']['optimizer']['lr']
    opt_G   = optim.Adam(list(G_AB.parameters()) + list(G_BA.parameters()), lr=lr, betas=(0.5, 0.999))
    opt_D_A = optim.Adam(D_A.parameters(), lr=lr, betas=(0.5, 0.999))
    opt_D_B = optim.Adam(D_B.parameters(), lr=lr, betas=(0.5, 0.999))

    cycle_loss = nn.L1Loss().to(device)

    # --------------------
    # Train
    # --------------------
    G_AB, G_BA, D_A, D_B, history = util_model.train_dualgan(
        G_AB, G_BA, D_A, D_B,
        data_loaders=data_loaders,
        opt_G=opt_G, opt_D_A=opt_D_A, opt_D_B=opt_D_B,
        cycle_loss=cycle_loss,
        cfg_trainer=cfg['trainer'],
        model_fold_dir=model_fold_dir,
        device=device
    )

    util_model.plot_training(history, model_name=model_name, plot_training_dir=plot_training_fold_dir)

    # --------------------
    # Test A->B
    # --------------------
    mse_t, psnr_t, vif_t, ssim_t = util_model.test_4metrics(
        G_AB, loader_test=data_loaders["test"], device=device,
        outputs_dir=outputs_fold_dir, save_output=True
    )
    print(f"Fold {fold} - MSE: {mse_t:.6f}, PSNR: {psnr_t:.4f}, VIF: {vif_t:.4f}, SSIM: {ssim_t:.4f}")

    results_mse[f"{fold} MSE"].append(mse_t)
    results_psnr[f"{fold} PSNR"].append(psnr_t)
    results_vif[f"{fold} VIF"].append(vif_t)
    results_ssim[f"{fold} SSIM"].append(ssim_t)

    # Save reports
    df_mse = pd.DataFrame(results_mse)
    df_mse.insert(0, "std MSE", df_mse[metric_cols_mse].std(axis=1))
    df_mse.insert(0, "mean MSE", df_mse[metric_cols_mse].mean(axis=1))
    df_mse.insert(0, "model", model_name)
    df_mse.to_excel(report_file_mse, index=False)

    df_psnr = pd.DataFrame(results_psnr)
    df_psnr.insert(0, "std PSNR", df_psnr[metric_cols_psnr].std(axis=1))
    df_psnr.insert(0, "mean PSNR", df_psnr[metric_cols_psnr].mean(axis=1))
    df_psnr.insert(0, "model", model_name)
    df_psnr.to_excel(report_file_psnr, index=False)

    df_vif = pd.DataFrame(results_vif)
    df_vif.insert(0, "std VIF", df_vif[metric_cols_vif].std(axis=1))
    df_vif.insert(0, "mean VIF", df_vif[metric_cols_vif].mean(axis=1))
    df_vif.insert(0, "model", model_name)
    df_vif.to_excel(report_file_vif, index=False)

    df_ssim = pd.DataFrame(results_ssim)
    df_ssim.insert(0, "std SSIM", df_ssim[metric_cols_ssim].std(axis=1))
    df_ssim.insert(0, "mean SSIM", df_ssim[metric_cols_ssim].mean(axis=1))
    df_ssim.insert(0, "model", model_name)
    df_ssim.to_excel(report_file_ssim, index=False)
